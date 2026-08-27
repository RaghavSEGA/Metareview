"""
Builds the metareview matrix xlsx — same structure as Sega's existing MetareviewMatrix
template (validated in the Persona 3: Reload POC). Pure function of (reviews, categories,
data) -> openpyxl Workbook, so it's easy to unit test without Streamlit or the network.
"""
from io import BytesIO

import openpyxl
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .charts import render_opinion_graph_png

FONT = "Arial"
FIRST_REVIEW_COL = 6  # column F, matching the original template layout

_COMMENT_WIDTH = 260  # wide enough that _COMMENT_CHARS_PER_LINE stays a fair estimate
_COMMENT_CHARS_PER_LINE = 34  # rough wrap width at that box's default comment font
_COMMENT_MIN_HEIGHT, _COMMENT_MAX_HEIGHT = 90, 400


def _comment_box_size(text: str):
    """
    openpyxl's Comment defaults to a fixed 144x79px box (openpyxl.comments.Comment's own
    defaults) — comfortably fits a short pull-quote but clips anything longer, so a producer
    only sees the box's first line or two unless they manually drag it bigger in Excel. The
    actual comment TEXT was never truncated (Comment.content always held the full string) — it
    just wasn't visible without resizing, which reads the same as truncation to anyone who
    didn't think to try. Sizing the box's height to the quote's real length up front means the
    full quote is visible on open, capped so one enormous quote can't cover half the sheet.
    """
    lines_needed = max(1, -(-len(text) // _COMMENT_CHARS_PER_LINE))  # ceil division
    height = min(_COMMENT_MAX_HEIGHT, max(_COMMENT_MIN_HEIGHT, lines_needed * 16 + 20))
    return _COMMENT_WIDTH, height


def build_matrix_workbook(game_title: str, reviews, categories, data) -> openpyxl.Workbook:
    """
    reviews: list of dicts with keys key, outlet, score, date
    categories: ordered list of category names (rows)
    data: {category: {review_key: (value, quote)}}
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    n = len(reviews)
    if n == 0:
        raise ValueError("Need at least one review to build a matrix")
    last_col_letter = get_column_letter(FIRST_REVIEW_COL + n - 1)
    first_col_letter = get_column_letter(FIRST_REVIEW_COL)

    ws["A1"] = f"{game_title} — Metareview"
    ws["A1"].font = Font(name=FONT, bold=True, size=14)

    ws["A2"] = "Review Count:"
    ws["A2"].font = Font(name=FONT, bold=True)
    # Count off row 2 (outlet names) — always populated for every included review — not row 5
    # (Score). Not every review states a numeric score (and shouldn't be forced to), so counting
    # off the Score row understated "Review Count" as 0 on a real run where every review's score
    # cell happened to be blank, even though 35 real reviews were plainly present in the sheet.
    ws["E2"] = f"=COUNTA({first_col_letter}2:{last_col_letter}2)"

    ws["A3"] = "Average Score:"
    ws["A3"].font = Font(name=FONT, bold=True)
    # IFERROR guards against #DIV/0! when literally no review in the batch has a stated score
    # (legitimate — plenty of outlets don't publish one) — shows a clear "N/A" instead of an
    # unexplained Excel error in a producer-facing report.
    ws["E3"] = f'=IFERROR(AVERAGE({first_col_letter}5:{last_col_letter}5), "N/A")'

    ws["A4"] = "Comparison (weighted)"
    ws["B4"], ws["C4"], ws["D4"], ws["E4"] = "Positive", "Mixed", "Negative", "Review Site"
    for c in ["A4", "B4", "C4", "D4", "E4"]:
        ws[c].font = Font(name=FONT, bold=True)

    ws["E5"] = "Score"
    ws["E6"] = "Date Posted"

    for i, r in enumerate(reviews):
        col = FIRST_REVIEW_COL + i
        ws.cell(row=2, column=col, value=r["outlet"]).font = Font(name=FONT, bold=True)
        ws.cell(row=5, column=col, value=r.get("score"))
        ws.cell(row=6, column=col, value=r.get("date"))

    pos_fill = PatternFill("solid", fgColor="C6EFCE")
    neg_fill = PatternFill("solid", fgColor="FFC7CE")
    neu_fill = PatternFill("solid", fgColor="FFEB9C")

    row = 7
    for cat in categories:
        # Same fix as Review Count above: the denominator here is the total number of reviews
        # in the batch (per the weighted-score formula: (positive-negative)/review_count), which
        # row 2 (outlet names) always reflects correctly — row 5 (Score) does not, since a score
        # is often missing for some or all reviews for entirely legitimate reasons.
        ws.cell(row=row, column=1,
                 value=f"=(B{row}-D{row})/(COUNTA(${first_col_letter}$2:${last_col_letter}$2))")
        ws.cell(row=row, column=2, value=f'=COUNTIF({first_col_letter}{row}:{last_col_letter}{row}, "1")')
        ws.cell(row=row, column=3, value=f'=COUNTIF({first_col_letter}{row}:{last_col_letter}{row}, "0")')
        ws.cell(row=row, column=4, value=f'=COUNTIF({first_col_letter}{row}:{last_col_letter}{row}, "-1")')
        ws.cell(row=row, column=5, value=cat)

        for i, r in enumerate(reviews):
            col = FIRST_REVIEW_COL + i
            entry = data.get(cat, {}).get(r["key"])
            if entry:
                val, quote = entry
                cell = ws.cell(row=row, column=col, value=val)
                width, height = _comment_box_size(quote)
                cell.comment = Comment(quote, "AI Metareview", width=width, height=height)
                cell.fill = pos_fill if val == 1 else (neg_fill if val == -1 else neu_fill)
        row += 1

    last_cat_row = row - 1
    # A rendered picture, not a native openpyxl BarChart — see charts.py's module docstring for
    # why: a native chart object here was confirmed to render with blank category-axis labels
    # (an empty, rotated placeholder box per label) in non-Excel viewers, even though the same
    # chart displays fine in Excel desktop. A picture looks identical everywhere.
    scores = compute_weighted_scores(reviews, categories, data)
    opinion_graph_png = render_opinion_graph_png(game_title, scores)
    if opinion_graph_png:
        img = XLImage(BytesIO(opinion_graph_png))
        ws.add_image(img, f"E{last_cat_row + 3}")

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["E"].width = 22
    for i in range(n):
        ws.column_dimensions[get_column_letter(FIRST_REVIEW_COL + i)].width = 14

    return wb


def compute_weighted_scores(reviews, categories, data, threshold=0.25):
    """Python-side mirror of the xlsx formulas, for on-screen preview/charting without
    needing to round-trip through Excel (and for unit tests)."""
    n = len(reviews)
    results = {}
    for cat in categories:
        entries = data.get(cat, {})
        pos = sum(1 for v, _ in entries.values() if v == 1)
        neu = sum(1 for v, _ in entries.values() if v == 0)
        neg = sum(1 for v, _ in entries.values() if v == -1)
        mentioned = pos + neu + neg
        results[cat] = {
            "weighted": (pos - neg) / n if n else 0,
            "positive": pos, "mixed": neu, "negative": neg,
            "mention_rate": mentioned / n if n else 0,
            "meets_threshold": (mentioned / n if n else 0) >= threshold,
            "suggested_label": suggest_sentiment_label(pos, neu, neg),
        }
    return results


def compute_platform_averages(reviews):
    """
    Per-platform breakdown of review count and average score — the report's "Review Averages"
    section shows this as extra lines under the overall blended number when a run has reviews
    tagged with 2+ distinct platforms, mirroring the per-storefront/platform breakdown seen in
    Sega's own Puyo Puyo Tetris reference report (that one split by storefront too — Metacritic,
    Amazon, GameStop — this only splits by platform, since that's the dimension this app
    actually has data for, mainly via the Metacritic all-platforms sourcing in app/metacritic.py).

    reviews: list of dicts with at least `score` and `platform`. `platform` may be None/blank —
    those reviews are simply excluded from this breakdown (not lumped into a fake "Unknown"
    bucket); the overall Review Averages line already covers every review regardless of whether
    it's platform-tagged.

    Returns {platform: {"review_count": n, "average_score": avg_or_None}}, sorted by review
    count descending (ties broken alphabetically) so the most-reviewed platform leads. A
    platform with reviews but none carrying a stated score still gets an entry (review_count > 0,
    average_score None) rather than silently vanishing — same "N/A, not zero" principle as the
    overall average.
    """
    by_platform = {}
    for r in reviews:
        platform = (r.get("platform") or "").strip()
        if not platform:
            continue
        by_platform.setdefault(platform, []).append(r.get("score"))

    result = {}
    for platform, scores in by_platform.items():
        scored = [s for s in scores if s is not None]
        result[platform] = {
            "review_count": len(scores),
            "average_score": round(sum(scored) / len(scored), 2) if scored else None,
        }
    return dict(sorted(result.items(), key=lambda kv: (-kv[1]["review_count"], kv[0])))


def suggest_sentiment_label(positive: int, mixed: int, negative: int) -> str:
    """
    A deterministic first guess at the severity label Sega's reports use in category
    call-outs (e.g. "Combining Puyo Puyo and Tetris [Widely Praised]"), derived from the
    Persona 3: Reload and Puyo Puyo Tetris reference reports. This is a STARTING POINT, not
    a verdict — the narrative-drafting prompt (see narrative.py) is told it may override this
    with a one-line rationale when a category's sample is thin, exactly the way the human
    producer's "Richard's Note" caveat overrides a label on a lightly-commented category in
    the Puyo Puyo Tetris report. Don't chase perfect agreement with every historical label;
    some of those were themselves a judgment call on a tiny sample.

    Rule of thumb, in order:
      - no negative, no mixed              -> "Universally Praised"   (or Panned if positive=0)
      - no negative, mixed is a minority    -> "Widely Praised"        (or Panned)
      - no negative, mixed is not a minority-> "Generally Praised"     (or Panned)
      - both positive and negative present, each a real share of the total -> "Extremely Controversial"
      - net positive, negative a small minority -> "Somewhat Controversial, Leaning Positive"
      - net positive, negative a real minority  -> "Controversial, Leaning Positive"
      - (mirrored for net negative)
    """
    total = positive + mixed + negative
    if total == 0 or (positive == 0 and negative == 0):
        return "Not enough data"

    pos_share, neg_share, mixed_share = positive / total, negative / total, mixed / total

    # A small minority of dissent (<=12% of mentions) doesn't disqualify "praised" — e.g.
    # Puyo Puyo Tetris's "Combining Puyo Puyo and Tetris" had 9 positive/1 negative out of 10
    # and the report still called it "Widely Praised", not controversial.
    if neg_share <= 0.12:
        if negative == 0 and mixed == 0:
            return "Universally Praised"
        return "Widely Praised" if mixed_share < 0.30 else "Generally Praised"
    if pos_share <= 0.12:
        if positive == 0 and mixed == 0:
            return "Universally Panned"
        return "Widely Panned" if mixed_share < 0.30 else "Generally Panned"

    minor_share = min(pos_share, neg_share)
    # "Extremely Controversial" needs real volume behind it, not just an unlucky small
    # sample (3 mentions split 1/2 is a thin lean, not a controversy) — either a substantial
    # minority on both sides, or no single bucket dominating at all (high mixed_share too).
    extremely_controversial = total >= 6 and (
        minor_share >= 0.30 or (minor_share >= 0.20 and mixed_share >= 0.35)
    )
    if extremely_controversial:
        return "Extremely Controversial"

    leaning = "Positive" if positive >= negative else "Negative"
    qualifier = "Somewhat Controversial" if minor_share < 0.25 else "Controversial"
    return f"{qualifier}, Leaning {leaning}"
